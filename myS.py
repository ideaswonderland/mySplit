from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Border, Side, Alignment
import pandas as pd
import numpy as np
import os

desktop = os.path.join(os.path.join(os.environ['USERPROFILE']), 'Desktop') 

class MYS():
    def __init__(self,
                 df_mys:pd.DataFrame,
                 df_mebbis:pd.DataFrame,
                 df_mebbis_dummy:pd.DataFrame,
                 df_kurum_ana:pd.DataFrame,
                 df_kurum_mem:pd.DataFrame,
                 firma:str,
                 faturaTür:str,
                 imzaListe:list,
                 dosyaYolu:str,
                 tekKaynak: dict,
                 harcamaTalimatı: dict
                 ):
        
        self.df_ana = pd.DataFrame() #boş dataframe ler oluşturuyor, düzeltilmiş dataframe i ayırıp birleştirecez
        self.df_ilk = pd.DataFrame()
        self.df_mem = pd.DataFrame()

        self.df_mys = df_mys
        self.df_mebbis = df_mebbis
        self.df_mebbis_dummy = df_mebbis_dummy
        self.df_kurum_ana = df_kurum_ana
        self.df_kurum_mem = df_kurum_mem
        self.firma = firma
        self.faturaTür = faturaTür
        self.imzaListe = imzaListe
        self.dosyaYolu = dosyaYolu
        self.tekKaynak = tekKaynak
        self.harcamaTalimatı = harcamaTalimatı

        self.meb_kur_col = df_mebbis.columns.get_loc('KURUM ADI')
        self.meb_abo_col = df_mebbis.columns.get_loc('ABONE NUMARASI')
        self.meb_ver_col = df_mebbis.columns.get_loc('VERGİ NO')
        self.meb_fat_col = df_mebbis.columns.get_loc('FATURA NUMARASI')
        self.meb_tar_col = df_mebbis.columns.get_loc('FATURA TARİHİ')
        self.meb_icm_col = df_mebbis.columns.get_loc('İCMAL NO')
        self.meb_tuk_col = df_mebbis.columns.get_loc('TÜKETİM MİKTARI')
        self.meb_tut_col = df_mebbis.columns.get_loc('FATURA TUTARI')
        
        self.mebD_ver_col = df_mebbis_dummy.columns.get_loc('VERGİ NO')
        self.mebD_icm_col = df_mebbis_dummy.columns.get_loc('İCMAL NO')
        
        self.mys_fat_col = df_mys.columns.get_loc('Fatura No')
        self.mys_tut_col = df_mys.columns.get_loc('Ödenecek Tutar')
        self.mys_tar_col = df_mys.columns.get_loc('Tarih')
        self.mys_ver_col = df_mys.columns.get_loc('VKN')

        yolTemp = r'%s\Taslaklar\temp.xlsx'%(dosyaYolu)
        
        self.wb = load_workbook(yolTemp)
        self.ws = self.wb['Liste']
        
    def MYS(self):
        wb = self.wb
        ws = self.ws
        tekKaynak = self.tekKaynak
        harcamaTalimatı = self.harcamaTalimatı

        for i in range(len(self.df_mebbis.index)): #baz olarak mebbis dosyasını alıp oradaki tüm satırları teker teker işleme alıyor
            indices_fatura = np.where(self.df_mys['Fatura No'] == self.df_mebbis.iat[i,self.meb_fat_col]) #iki dosyadaki fatura numaralarını eşleştiriyor
            row_indices_fatura = indices_fatura[0] #faturaların mys dosyasındaki konumunu buluyor
            if row_indices_fatura.size == 0: #Fatura numarasının mys dosyasında olup olmadığını kontrol ediyor
                aboneNo = self.df_mys['Abone'].values == self.df_mebbis.iat[i,self.meb_abo_col] #Fatura numarası mys dosyasında yoksa abone numarasını buluyor.
                tarKontrol = self.df_mys['Tarih'].values == self.df_mebbis.iat[i,self.meb_tar_col]
                if max(aboneNo)==True and max(tarKontrol)==True: #abone numarasının mys dosyasında olup olmadığını kontrol ediyor
                    row_indices_df_mys = np.where((aboneNo == True)&(tarKontrol == True)) #Abone numarasının hangi sırada olduğunu buluyor
                    #Tutar ve fatura tarihi kontrolü
                    if self.df_mebbis.iat[i,self.meb_tut_col] == self.df_mys.iat[row_indices_df_mys[0][0],self.mys_tut_col] and self.df_mebbis.iat[i,self.meb_tar_col] == self.df_mys.iat[row_indices_df_mys[0][0],self.mys_tar_col]:
                        #Tutar ve tarih iki tabloda da aynı ise mys'deki fatura numarasını alıyor
                        self.df_mebbis.iat[i,self.meb_fat_col] = self.df_mys.iat[row_indices_df_mys[0][0],self.mys_fat_col]                          
                    else: #tutar ve tarih tutmuyorsa fatura girişi hatalı demektir 
                        print(f'{self.df_mebbis.iat[i,self.meb_fat_col]} Fatura girişi hatalı!')
                else: #abone numarası yoksa zaten mys'de yok demektir
                    print(f'Fatura listesinde {i+2}. satırda yer alan {self.df_mebbis.iat[i,self.meb_fat_col]} MYS de yok!')                      
            else:
                #fatura var ancak tutarlar farklı ise tutarı mys'den çekiyor
                if self.df_mebbis.iat[i, self.meb_tut_col] != self.df_mys.iat[row_indices_fatura[0],self.mys_tut_col]:
                    self.df_mebbis.iat[i, self.meb_tut_col] = self.df_mys.iat[row_indices_fatura[0],self.mys_tut_col]
                    #tutar değişimi olursa uyarı veriyor, hangi faturanın tutarı nasıl değişmiş görebiliyoruz
                    print(f'{self.df_mebbis.iat[i, self.meb_fat_col]} numaralı fatura tutarı {self.df_mebbis.iat[i, self.meb_tut_col]} iken {self.df_mys.iat[row_indices_fatura[0],self.mys_tut_col]} oldu.')
            #vergi numarasını metin haline getiriyor böylece başta 0 varsa yoksaymıyor
            self.df_mebbis.iat[i,self.meb_ver_col] = self.df_mebbis_dummy.iat[i,self.mebD_ver_col]
            self.df_mebbis.iat[i,self.meb_icm_col] = self.df_mebbis_dummy.iat[i,self.mebD_icm_col]
            #kurum anaokulları içerisinde mi onu kontrol ediyor 
            if self.df_kurum_ana.isin([self.df_mebbis.iat[i,self.meb_ver_col]]).any().any():
                #anaokulları içerisindeyse anaokulu dataframe ine ekliyor
                self.df_ana = pd.concat([self.df_ana,self.df_mebbis.iloc[[i]]])
            elif self.df_kurum_mem.isin([self.df_mebbis.iat[i,self.meb_ver_col]]).any().any():
                self.df_mem = pd.concat([self.df_mem,self.df_mebbis.iloc[[i]]])
            else:
                #değilse ilkokullar dataframe ine ekliyor
                self.df_ilk = pd.concat([self.df_ilk,self.df_mebbis.iloc[[i]]])

        ws.cell(row=2,column=2).value = self.faturaTür # type: ignore
        Hizala.Solda(self.wb,2,2)
        ws.cell(row=3,column=2).value = self.firma # type: ignore
        Hizala.Solda(wb,3,2)

        başlangıç = 5

        if self.df_mem.empty:
            toplam_mem = 0
            if self.df_ilk.empty:
                toplam_ilk = 0
                if self.df_ana.empty:
                    toplam_ana = 0
                    pass
                else:
                    toplam_ana = self.df_ana['FATURA TUTARI'].sum()
                    self.işle(wb,self.df_ana,başlangıç)
                    başlangıç += len(self.df_ana)
                    ws['L6'] = f'=SUM(I{başlangıç-len(self.df_ana)}:I{başlangıç-1})' ####
                    ws.cell(row=başlangıç, column=1).value = 'ANAOKULLARI TOPLAMI' # type: ignore
                    Hizala.SağdaKalın(wb,başlangıç,1)
                    ws.merge_cells(f'A{başlangıç}:G{başlangıç}')
                    ws.cell(row=başlangıç, column=self.meb_tut_col+1).value = toplam_ana # type: ignore
                    Hizala.SağdaKalın(wb,başlangıç,self.meb_tut_col+1)    # type: ignore
                    ws.cell(row=başlangıç, column=self.meb_tut_col+1).number_format = '#,##0.00' # type: ignore
                    ws['M6'] = f'=H{başlangıç}-L6' ###     
                    başlangıç += 1
            else:
                toplam_ilk = self.df_ilk['FATURA TUTARI'].sum()
                self.işle(wb, self.df_ilk,başlangıç)
                başlangıç += len(self.df_ilk)
                ws['L5'] = f'=SUM(I5:I{başlangıç-1})' ####
                
                ws.cell(row=başlangıç, column=1).value = 'İLKÖĞRETİM TOPLAMI' # type: ignore
                Hizala.SağdaKalın(wb,başlangıç,1)
                ws.merge_cells(f'A{başlangıç}:G{başlangıç}')
                ws.cell(row=başlangıç, column=self.meb_tut_col+1).value = toplam_ilk # type: ignore
                Hizala.SağdaKalın(wb,başlangıç,self.meb_tut_col+1) # type: ignore
                ws.cell(row=başlangıç, column=self.meb_tut_col+1).number_format = '#,##0.00' # type: ignore
                ws['M5'] = f'=H{başlangıç}-L5' ###
                başlangıç += 1
                
                if self.df_ana.empty:
                    toplam_ana = 0
                else:
                    toplam_ana = self.df_ana['FATURA TUTARI'].sum()
                    self.işle(wb,self.df_ana,başlangıç)
                    başlangıç +=len(self.df_ana)
                    ws['L6'] = f'=SUM(I{başlangıç-len(self.df_ana)}:I{başlangıç-1})' ####
                    ws.cell(row=başlangıç, column=1).value = 'ANAOKULLARI TOPLAMI' # type: ignore
                    Hizala.SağdaKalın(wb,başlangıç,1)
                    ws.merge_cells(f'A{başlangıç}:G{başlangıç}')
                    ws.cell(row=başlangıç, column=self.meb_tut_col+1).value = toplam_ana # type: ignore
                    Hizala.SağdaKalın(wb,başlangıç,self.meb_tut_col+1)    # type: ignore
                    ws.cell(row=başlangıç, column=self.meb_tut_col+1).number_format = '#,##0.00' # type: ignore
                    ws['M6'] = f'=H{başlangıç}-L6' ###             
                    başlangıç +=1
        else:
            toplam_ilk = 0
            toplam_ana = 0
            toplam_mem = self.df_mem['FATURA TUTARI'].sum()
            self.işle(wb,self.df_mem,başlangıç)
            başlangıç +=len(self.df_mem)
            ws['A1'] = f'SULTANHİSAR İLÇE MİLLİ EĞİTİM MÜDÜRLÜĞÜ FATURA LİSTESİ'
            ws['L6'] = f'=SUM(I{başlangıç-len(self.df_mem)}:I{başlangıç-1})' ####
            ws.cell(row=başlangıç, column=1).value = 'İLÇE MİLLİ EĞİTİM MÜDÜRLÜĞÜ TOPLAMI' # type: ignore
            Hizala.SağdaKalın(wb,başlangıç,1)
            ws.merge_cells(f'A{başlangıç}:G{başlangıç}')
            ws.cell(row=başlangıç, column=self.meb_tut_col+1).value = toplam_mem # type: ignore
            Hizala.SağdaKalın(wb,başlangıç,self.meb_tut_col+1)    # type: ignore
            ws.cell(row=başlangıç, column=self.meb_tut_col+1).number_format = '#,##0.00' # type: ignore
            ws['M6'] = f'=H{başlangıç}-L6' ###             
            başlangıç +=1
            
                
        ws.cell(row=başlangıç, column=1).value = 'GENEL TOPLAM' # type: ignore
        Hizala.SağdaKalın(wb,başlangıç,1)
        ws.merge_cells(f'A{başlangıç}:G{başlangıç}')      
        ws.cell(row=başlangıç, column=self.meb_tut_col+1).value = toplam_ilk + toplam_ana + toplam_mem # type: ignore
        Hizala.SağdaKalın(wb,başlangıç,self.meb_tut_col+1)    # type: ignore
        ws.cell(row=başlangıç, column=self.meb_tut_col+1).number_format = '#,##0.00'    # type: ignore
        başlangıç += 4

        ws.cell(row=başlangıç, column=1).value = self.imzaListe[0]
        Hizala.SadeOrta(wb,başlangıç,1)
        ws.cell(row=başlangıç+1, column=1).value = self.imzaListe[1]
        Hizala.SadeOrta(wb,başlangıç+1,1)
        ws.cell(row=başlangıç, column=7).value = self.imzaListe[2]
        Hizala.SadeOrta(wb,başlangıç,7)
        ws.cell(row=başlangıç+1, column=7).value = self.imzaListe[3]
        Hizala.SadeOrta(wb,başlangıç+1,7)

        ws.print_area = f'A1:H{başlangıç+3}'

        wb.save(r'%s\Fatura_%s.xlsx'%(desktop,self.faturaTür))

        wbTek = load_workbook(r'%s\Fatura_%s.xlsx'%(desktop,self.faturaTür))
        wsTek = wbTek['Tek']

        wsTek['C5'] = tekKaynak['firma']
        wsTek['C6'] = tekKaynak['tebligat']
        wsTek['C7'] = tekKaynak['vergi']
        wsTek['C8'] = tekKaynak['telefon']
        wsTek['C9'] = tekKaynak['eposta']
        wsTek['C12'] = tekKaynak['tutar']
        wsTek['A15'] = tekKaynak['ihtiyaç']
        wsTek['C40'] = tekKaynak['harcama']
        wsTek['C41'] = tekKaynak['unvan']

        wbTek.save(r'%s\Fatura_%s.xlsx'%(desktop,self.faturaTür))

        wbHar = load_workbook(r'%s\Fatura_%s.xlsx'%(desktop,self.faturaTür))
        wsHar = wbHar['Harcama']
    
        ilk = toplam_ilk
        ana = toplam_ana
        mem = toplam_mem
        
        ilkT = '{:,.2f}'.format(ilk)
        ilkT = ilkT.replace(',','-')
        ilkT = ilkT.replace('.',',')
        ilkT = ilkT.replace('-','.')        

        anaT = '{:,.2f}'.format(ana)
        anaT = anaT.replace(',','-')
        anaT = anaT.replace('.',',')
        anaT = anaT.replace('-','.')
        
        memT = '{:,.2f}'.format(mem)
        memT = memT.replace(',','-')
        memT = memT.replace('.',',')
        memT = memT.replace('-','.')
        
        ilkTertip = harcamaTalimatı['ilkTertip']
        anaTertip = harcamaTalimatı['anaTertip']
        memTertip = harcamaTalimatı['memTertip']
        
        ilkS = f'{ilkTertip} ({ilkT} TL)'
        anaS = f'{anaTertip} ({anaT} TL)'
        memS = f'{memTertip} ({memT} TL)'
        
        if mem == 0:
            if ilk == 0:
                ilkS = anaS
                anaS = ''
            if ana == 0:
                anaS = ''
        else:
            ilkS = memS
            anaS=''
            
        self.harcamaTalimatı['ödenek1'] = ilkS
        self.harcamaTalimatı['ödenek2'] = anaS
        
        wsHar['C4'] = harcamaTalimatı['tarih']
        wsHar['C8'] = harcamaTalimatı['tanım']
        wsHar['C9'] = harcamaTalimatı['nitelik']
        wsHar['C10'] = harcamaTalimatı['miktar']
        wsHar['C11'] = harcamaTalimatı['ödenek1']
        wsHar['C12'] = harcamaTalimatı['ödenek2']
        wsHar['A16'] = harcamaTalimatı['metin']
        wsHar['A23'] = self.imzaListe[0]
        wsHar['A24'] = self.imzaListe[1]
        wsHar['D23'] = self.imzaListe[2]
        wsHar['D24'] = self.imzaListe[3]  
        
        wbHar.save(r'%s\Fatura_%s.xlsx'%(desktop,self.faturaTür))    

    def işle(self, wb:Workbook, df:pd.DataFrame, başlangıç:int):
        ws = wb.active
        for i in range(len(df.values)):
            ws.cell(row=başlangıç+i, column=self.meb_kur_col+1).value = df.values[i][self.meb_kur_col] # type: ignore
            Hizala.Solda(wb,başlangıç+i,self.meb_kur_col+1) # type: ignore
            ws.cell(row=başlangıç+i, column=self.meb_abo_col+1).value = df.values[i][self.meb_abo_col] # type: ignore
            Hizala.Ortala(wb,başlangıç+i,self.meb_abo_col+1) # type: ignore
            ws.cell(row=başlangıç+i, column=self.meb_ver_col+1).value = df.values[i][self.meb_ver_col] # type: ignore
            Hizala.Ortala(wb,başlangıç+i,self.meb_ver_col+1) # type: ignore
            ws.cell(row=başlangıç+i, column=self.meb_fat_col+1).value = df.values[i][self.meb_fat_col] # type: ignore
            Hizala.Ortala(wb,başlangıç+i,self.meb_fat_col+1) # type: ignore
            ws.cell(row=başlangıç+i, column=self.meb_tar_col+1).value = df.values[i][self.meb_tar_col] # type: ignore
            Hizala.Ortala(wb,başlangıç+i,self.meb_tar_col+1) # type: ignore
            ws.cell(row=başlangıç+i, column=self.meb_icm_col+1).value = df.values[i][self.meb_icm_col] # type: ignore
            Hizala.Ortala(wb,başlangıç+i,self.meb_icm_col+1) # type: ignore
            ws.cell(row=başlangıç+i, column=self.meb_tuk_col+1).value = df.values[i][self.meb_tuk_col] # type: ignore
            Hizala.Ortala(wb,başlangıç+i,self.meb_tuk_col+1) # type: ignore
            ws.cell(row=başlangıç+i, column=self.meb_tut_col+1).value = df.values[i][self.meb_tut_col] # type: ignore
            ws.cell(row=başlangıç+i, column=self.meb_tut_col+1).number_format = '#,##0.00' # type: ignore
            Hizala.Sağda(wb,başlangıç+i,self.meb_tut_col+1)       # type: ignore
 
class Hizala():
    @staticmethod
    def Solda(wb, row, column):
        ws = wb.active
        ft = Font(color='FF000000', name='Times New Roman', size=10)
        thin = Side(border_style="thin", color="FF000000")
        
        ws.cell(row=row, column=column).font = ft
        ws.cell(row=row, column=column).alignment = Alignment(horizontal="left", vertical="center")
        ws.cell(row=row, column=column).border = Border(top=thin, left=thin, right=thin, bottom=thin)

    @staticmethod
    def Ortala(wb, row, column):
        ws = wb.active
        ft = Font(color='FF000000', name='Times New Roman', size=10)
        thin = Side(border_style="thin", color="FF000000")
        
        ws.cell(row=row, column=column).font = ft
        ws.cell(row=row, column=column).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=row, column=column).border = Border(top=thin, left=thin, right=thin, bottom=thin)

    @staticmethod
    def SadeOrta(wb, row, column):
        ws = wb.active
        ft = Font(color='FF000000', name='Times New Roman', size=10)
        
        ws.cell(row=row, column=column).font = ft
        ws.cell(row=row, column=column).alignment = Alignment(horizontal="center", vertical="center")

    @staticmethod
    def Sağda(wb, row, column):
        ws = wb.active
        ft = Font(color='FF000000', name='Times New Roman', size=10)
        thin = Side(border_style="thin", color="FF000000")
        
        ws.cell(row=row, column=column).font = ft
        ws.cell(row=row, column=column).alignment = Alignment(horizontal="right", vertical="center")
        ws.cell(row=row, column=column).border = Border(top=thin, left=thin, right=thin, bottom=thin)

    @staticmethod
    def SağdaKalın(wb, row, column):
        ws = wb.active
        ft = Font(color='FF000000', name='Times New Roman', size=12, bold=True)
        thin = Side(border_style="thin", color="FF000000")
        
        ws.cell(row=row, column=column).font = ft
        ws.cell(row=row, column=column).alignment = Alignment(horizontal="right", vertical="center")
        ws.cell(row=row, column=column).border = Border(top=thin, left=thin, right=thin, bottom=thin)