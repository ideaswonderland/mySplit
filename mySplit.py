from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Color, Border, Side, Alignment
import pandas as pd
import numpy as np
import os
import warnings
from paracevir import ParaCevir
from datetime import date

f = open("dosyaYolu.txt", "r")
dosyaYolu = r'%s'%(f.read())
f.close()
desktop = os.path.join(os.path.join(os.environ['USERPROFILE']), 'Desktop') 

class Fatura():
    def __init__(self,ay,yıl):
        
        #Genel dosya yolları
        self.dönem = r'%s/%s'%(yıl,ay)
        self.yolKurum = r'kurum.xlsx'
        self.yolİmza = r'imza.xlsx'
        #Elektrik dosya yolları
        self.yolFatura_E = r'%s\Elektrik\Fatura Listesi.xlsx'%(dosyaYolu)
        self.yolMYS_E = r'%s\Elektrik\MYS.xlsx'%(dosyaYolu)
        #İnternet dosya yolları
        self.yolFatura_I = r'%s\İnternet\Fatura Listesi.xlsx'%(dosyaYolu)
        self.yolMYS_I = r'%s\İnternet\MYS.xlsx'%(dosyaYolu)
        #Telefon dosya yolları
        self.yolFatura_T = r'%s\Telefon\Fatura Listesi.xlsx'%(dosyaYolu)
        self.yolMYS_T = r'%s\Telefon\MYS.xlsx'%(dosyaYolu)  
        #Doğalgaz dosya yolları      
        self.yolFatura_D = r'%s\Doğalgaz\Fatura Listesi.xlsx'%(dosyaYolu)
        self.yolMYS_D = r'%s\Doğalgaz\MYS.xlsx'%(dosyaYolu)  


    def Elektrik(self):
        yolFatura = self.yolFatura_E
        yolMYS = self.yolMYS_E
        yolKurum = self.yolKurum
        yolİmza = self.yolİmza
        
        df_firma = pd.read_excel(io='firma.xlsx',sheet_name='Elektrik', header=0)

        firma = df_firma.values[0][1]
        faturaTür = 'Elektrik Aboneliği Ödemesi'

        warnings.simplefilter(action='ignore', category=UserWarning)
        
        if os.path.exists(yolFatura):
            if os.path.exists(yolMYS):
                df_imza = pd.read_excel(yolİmza, header=0) #mys'den indirilen dosyayı okuyor
                imzaListe = [
                    df_imza.values[0][0],
                    df_imza.values[1][0],
                    df_imza.values[0][1],
                    df_imza.values[1][1]
                ]

                df_mys = pd.read_excel(yolMYS, header=0) #mys'den indirilen dosyayı okuyor
                cols_to_keep_mys = [
                    'Fatura No',
                    'Harcama Birimi',
                    'Fatura Tarihi',
                    'Ödenecek Tutar',
                    'Müşteri Kimlik Bilgisi'
                ]
                df_mys = df_mys.reindex(cols_to_keep_mys, axis=1) #mys dosyasından verileri çekiyor
                df_mys[['VKN', 'Okul']] = df_mys['Harcama Birimi'].str.split('-', expand=True) #Harcama Birimi kısmını VKN ve Okul olarak ayırıyor
                df_mys[['Tarih', 'Boş']] = df_mys['Fatura Tarihi'].str.split(' ', expand=True) #Fatura tarihi kısımını ayırıyor
                df_mys['Tarih'] = df_mys['Tarih'].str.replace('-', '/') #Fatura tarihini mebbis'ten alınan dosyanın formatına dönüştürüyor
                df_mys[['Sayaç No','Müşteri No','Tesisat No','Abone']]=df_mys['Müşteri Kimlik Bilgisi'].str.split('-', expand=True) #Müşteri kimlik bilgisi kısmını bölüyor
                df_mys['Abone']=df_mys['Abone'].astype(int) #abone numaralarını mebbis dosyasıyla eşlemek için tam sayı haline getiriyor
                #gereksiz sütunları temizliyor
                df_mys = df_mys.drop(columns=['Harcama Birimi',
                                            'Fatura Tarihi',
                                            'Boş',
                                            'Müşteri Kimlik Bilgisi',
                                            'Sayaç No',
                                            'Müşteri No',
                                            'Tesisat No',
                                            'Okul'])
                df_mys = df_mys[['Fatura No','Ödenecek Tutar','Tarih','Abone','VKN']]

                df_mebbis = pd.read_excel(yolFatura, header=0) #mebbis dosyasından verileri çekiyor
                cols_to_keep_meb = [
                    'KURUM ADI',
                    'ABONE NUMARASI', 
                    'VERGİ NO', 
                    'FATURA NUMARASI', 
                    'FATURA TARİHİ',
                    'İCMAL NO',
                    'TÜKETİM MİKTARI',
                    'FATURA TUTARI'
                ]
                df_mebbis = df_mebbis.reindex(cols_to_keep_meb, axis=1)
                df_mebbis = df_mebbis.drop(df_mebbis.index[-1]) #tablodaki son satırı siliyor
                df_mebbis['ABONE NUMARASI']=df_mebbis['ABONE NUMARASI'].astype(int) #abone numaralarını mys dosyasıyla eşlemek için tam sayı haline getiriyor
                df_mebbis['VERGİ NO']=df_mebbis['VERGİ NO'].astype(str) #ilerde hata almamak için kolonun veri tipini değiştiriyor
                df_mebbis['İCMAL NO']=df_mebbis['İCMAL NO'].astype(str) #böylece normalde sayısal veriler kayıpsız şekilde metin olarak saklanabiliyor
                
                df_mebbis_dummy = pd.read_excel(yolFatura, header=0, dtype=str) #aynı mebbis dosyasını okuyor ancak vergi no ve icmal metin olarak saklanmalı
                cols_to_keep_dummy = ['VERGİ NO','İCMAL NO'] #bu veriler sayısal olarak saklanırsa başında 0 varsa siliyor
                df_mebbis_dummy = df_mebbis_dummy.reindex(cols_to_keep_dummy, axis=1)
                df_mebbis_dummy = df_mebbis_dummy.drop(df_mebbis_dummy.index[-1])    

                df_kurum = pd.read_excel(yolKurum, dtype=str, header=0) #kurumlar dosyasından verileri çekiyor
                cols_to_keep_kurum = [
                            'VERGİ KİMLİK NO', 
                            'KURUM TÜRÜ'
                        ]
                df_kurum = df_kurum.reindex(cols_to_keep_kurum, axis=1)
                df_kurum =df_kurum.drop(df_kurum.index[-1])
                df_kurum_ana = df_kurum[df_kurum['KURUM TÜRÜ']=='Okul Öncesi'] #anaokullarını ayırıyor
                
                #Tek kaynak formu işlemleri
                toplam = "{:,.2f}".format(df_mebbis['FATURA TUTARI'].sum())
                toplam = toplam.replace(',','-')
                toplam = toplam.replace('.',',')
                toplam = toplam.replace('-','.')
                top2 = df_mebbis['FATURA TUTARI'].sum()
                tutar = f'{toplam} TL ({ParaCevir(top2)})'
                unik = len(df_mebbis['KURUM ADI'].unique())
                if unik<4:
                    kurumlar = ", ".join(str(element) for element in df_mebbis['KURUM ADI'].unique())
                else:
                    kurumlar = f'{len(df_mebbis)} Adet Fatura'
                icmaller = ", ".join(str(element) for element in df_mebbis_dummy['İCMAL NO'].unique())
                ihtiyaç = f'TEMEL EĞİTİM OKULLARI {self.dönem} DÖNEM ELEKTRİK ÖDEMESİ {kurumlar} (İcmal No: {icmaller})'
                
                
                tekKaynak = {
                    'firma' : df_firma.values[0][1],
                    'tebligat' : df_firma.values[1][1],
                    'vergi' : df_firma.values[2][1],
                    'telefon' : df_firma.values[3][1],
                    'eposta' : df_firma.values[4][1],
                    'tutar' : tutar,
                    'ihtiyaç' : ihtiyaç,
                    'harcama' : imzaListe[2],
                    'unvan' : imzaListe[3]
                }
                
                tür = 'elektrik'
                ilkTertip = f'40.149.423.18734.13.68.01.03.02'
                anaTertip = f'40.149.422.18735.13.68.01.03.02'
                nitelik = f'Temel Eğitim Okulları {self.dönem} Dönem Elektrik Ödemesi' ###
                metin = f'      Müdürlüğümüz  Temel Eğitim Okullarının ({kurumlar}) {self.dönem} dönem {tür} aboneliklerine ait toplam {tutar} borç ödemesi hususunu Onaylarınıza arz ederim.'
                
                harcamaTalimatı = {
                    'tarih' : date.today().strftime("%d.%m.%Y"),
                    'tanım' : faturaTür,
                    'nitelik' : nitelik,
                    'miktar' : tutar,
                    'ödenek1' : '',
                    'ödenek2' : '',
                    'metin' : metin,
                    'ilkTertip': ilkTertip,
                    'anaTertip': anaTertip
                }                
                

                MYS(df_mys,df_mebbis,df_mebbis_dummy, df_kurum_ana, firma, faturaTür, imzaListe, dosyaYolu, tekKaynak, harcamaTalimatı).MYS()
                
    def İnternet(self):
        yolFatura = self.yolFatura_I
        yolMYS = self.yolMYS_I
        yolKurum = self.yolKurum
        yolİmza = self.yolİmza

        df_firma = pd.read_excel(io='firma.xlsx',sheet_name='İnternet', header=0)

        firma = df_firma.values[0][1]
        faturaTür = 'İnternet Aboneliği Ödemesi'

        warnings.simplefilter(action='ignore', category=UserWarning)
        
        if os.path.exists(yolFatura):
            if os.path.exists(yolMYS):
                df_imza = pd.read_excel(yolİmza, header=0) #mys'den indirilen dosyayı okuyor
                imzaListe = [
                    df_imza.values[0][0],
                    df_imza.values[1][0],
                    df_imza.values[0][1],
                    df_imza.values[1][1]
                ]

                df_mys = pd.read_excel(yolMYS, header=0) #mys'den indirilen dosyayı okuyor
                cols_to_keep_mys = [
                    'Fatura No',
                    'Harcama Birimi',
                    'Fatura Tarihi',
                    'Ödenecek Tutar',
                    'Müşteri Kimlik Bilgisi'
                ]
                df_mys = df_mys.reindex(cols_to_keep_mys, axis=1) #mys dosyasından verileri çekiyor
                df_mys[['VKN', 'Okul']] = df_mys['Harcama Birimi'].str.split('-', expand=True) #Harcama Birimi kısmını VKN ve Okul olarak ayırıyor
                df_mys[['Tarih', 'Boş']] = df_mys['Fatura Tarihi'].str.split(' ', expand=True) #Fatura tarihi kısımını ayırıyor
                df_mys['Tarih'] = df_mys['Tarih'].str.replace('-', '/') #Fatura tarihini mebbis'ten alınan dosyanın formatına dönüştürüyor
                df_mys[['Hizmet No','Abone']]=df_mys['Müşteri Kimlik Bilgisi'].str.split('-', expand=True) #Müşteri kimlik bilgisi kısmını bölüyor
                df_mys['Abone']=df_mys['Abone'].astype(float) #abone numaralarını mebbis dosyasıyla eşlemek için tam sayı haline getiriyor
                #gereksiz sütunları temizliyor
                df_mys = df_mys.drop(columns=['Harcama Birimi',
                                            'Fatura Tarihi',
                                            'Boş',
                                            'Müşteri Kimlik Bilgisi',
                                            'Hizmet No',
                                            'Okul'])
                df_mys = df_mys[['Fatura No','Ödenecek Tutar','Tarih','Abone','VKN']]

                df_mebbis = pd.read_excel(yolFatura, header=0) #mebbis dosyasından verileri çekiyor
                cols_to_keep_meb = [
                    'KURUM ADI',
                    'ABONE NUMARASI', 
                    'VERGİ NO', 
                    'FATURA NUMARASI', 
                    'FATURA TARİHİ',
                    'İCMAL NO',
                    'TÜKETİM MİKTARI',
                    'FATURA TUTARI'
                ]
                df_mebbis = df_mebbis.reindex(cols_to_keep_meb, axis=1)
                df_mebbis = df_mebbis.drop(df_mebbis.index[-1]) #tablodaki son satırı siliyor
                df_mebbis['ABONE NUMARASI']=df_mebbis['ABONE NUMARASI'].astype(int) #abone numaralarını mys dosyasıyla eşlemek için tam sayı haline getiriyor
                df_mebbis['VERGİ NO']=df_mebbis['VERGİ NO'].astype(str) #ilerde hata almamak için kolonun veri tipini değiştiriyor
                df_mebbis['İCMAL NO']=df_mebbis['İCMAL NO'].astype(str) #böylece normalde sayısal veriler kayıpsız şekilde metin olarak saklanabiliyor
                
                df_mebbis_dummy = pd.read_excel(yolFatura, header=0, dtype=str) #aynı mebbis dosyasını okuyor ancak vergi no ve icmal metin olarak saklanmalı
                cols_to_keep_dummy = ['VERGİ NO','İCMAL NO'] #bu veriler sayısal olarak saklanırsa başında 0 varsa siliyor
                df_mebbis_dummy = df_mebbis_dummy.reindex(cols_to_keep_dummy, axis=1)
                df_mebbis_dummy = df_mebbis_dummy.drop(df_mebbis_dummy.index[-1])    

                df_kurum = pd.read_excel(yolKurum, dtype=str, header=0) #kurumlar dosyasından verileri çekiyor
                cols_to_keep_kurum = [
                            'VERGİ KİMLİK NO', 
                            'KURUM TÜRÜ'
                        ]
                df_kurum = df_kurum.reindex(cols_to_keep_kurum, axis=1)
                df_kurum =df_kurum.drop(df_kurum.index[-1])
                df_kurum_ana = df_kurum[df_kurum['KURUM TÜRÜ']=='Okul Öncesi'] #anaokullarını ayırıyor
                
                #Tek kaynak formu işlemleri
                toplam = "{:,.2f}".format(df_mebbis['FATURA TUTARI'].sum())
                toplam = toplam.replace(',','-')
                toplam = toplam.replace('.',',')
                toplam = toplam.replace('-','.')
                top2 = df_mebbis['FATURA TUTARI'].sum()
                tutar = f'{toplam} TL ({ParaCevir(top2)})'
                unik = len(df_mebbis['KURUM ADI'].unique())
                if unik<4:
                    kurumlar = ", ".join(str(element) for element in df_mebbis['KURUM ADI'].unique())
                else:
                    kurumlar = f'{len(df_mebbis)} Adet Fatura'
                icmaller = ", ".join(str(element) for element in df_mebbis_dummy['İCMAL NO'].unique())
                ihtiyaç = f'TEMEL EĞİTİM OKULLARI {self.dönem} DÖNEM İNTERNET ÖDEMESİ {kurumlar} (İcmal No: {icmaller})'
                tekKaynak = {
                    'firma' : df_firma.values[0][1],
                    'tebligat' : df_firma.values[1][1],
                    'vergi' : df_firma.values[2][1],
                    'telefon' : df_firma.values[3][1],
                    'eposta' : df_firma.values[4][1],
                    'tutar' : tutar,
                    'ihtiyaç' : ihtiyaç,
                    'harcama' : imzaListe[2],
                    'unvan' : imzaListe[3]
                }
                
                tür = 'internet'
                ilkTertip = f'40.149.423.292.13.68.01.03.05'
                anaTertip = f'40.149.422.291.13.68.01.03.05'
                nitelik = f'Temel Eğitim Okulları {self.dönem} Dönem İnternet Ödemesi' ###
                metin = f'      Müdürlüğümüz  Temel Eğitim Okullarının ({kurumlar}) {self.dönem} dönem {tür} aboneliklerine ait toplam {tutar} borç ödemesi hususunu Onaylarınıza arz ederim.'
                
                harcamaTalimatı = {
                    'tarih' : date.today().strftime("%d.%m.%Y"),
                    'tanım' : faturaTür,
                    'nitelik' : nitelik,
                    'miktar' : tutar,
                    'ödenek1' : '',
                    'ödenek2' : '',
                    'metin' : metin,
                    'ilkTertip': ilkTertip,
                    'anaTertip': anaTertip
                }       

                MYS(df_mys,df_mebbis,df_mebbis_dummy, df_kurum_ana, firma, faturaTür, imzaListe, dosyaYolu, tekKaynak, harcamaTalimatı).MYS()

    def Telefon(self):
        yolFatura = self.yolFatura_T
        yolMYS = self.yolMYS_T
        yolKurum = self.yolKurum
        yolİmza = self.yolİmza

        df_firma = pd.read_excel(io='firma.xlsx',sheet_name='Telefon', header=0)

        firma = df_firma.values[0][1]
        faturaTür = 'Telefon Aboneliği Ödemesi'

        warnings.simplefilter(action='ignore', category=UserWarning)
        
        if os.path.exists(yolFatura):
            if os.path.exists(yolMYS):
                df_imza = pd.read_excel(yolİmza, header=0) #mys'den indirilen dosyayı okuyor
                imzaListe = [
                    df_imza.values[0][0],
                    df_imza.values[1][0],
                    df_imza.values[0][1],
                    df_imza.values[1][1]
                ]

                df_mys = pd.read_excel(yolMYS, header=0) #mys'den indirilen dosyayı okuyor
                cols_to_keep_mys = [
                    'Fatura No',
                    'Harcama Birimi',
                    'Fatura Tarihi',
                    'Ödenecek Tutar',
                    'Müşteri Kimlik Bilgisi'
                ]
                df_mys = df_mys.reindex(cols_to_keep_mys, axis=1) #mys dosyasından verileri çekiyor
                df_mys[['VKN', 'Okul']] = df_mys['Harcama Birimi'].str.split('-', expand=True) #Harcama Birimi kısmını VKN ve Okul olarak ayırıyor
                df_mys[['Tarih', 'Boş']] = df_mys['Fatura Tarihi'].str.split(' ', expand=True) #Fatura tarihi kısımını ayırıyor
                df_mys['Tarih'] = df_mys['Tarih'].str.replace('-', '/') #Fatura tarihini mebbis'ten alınan dosyanın formatına dönüştürüyor
                df_mys[['Hizmet No','Abone']]=df_mys['Müşteri Kimlik Bilgisi'].str.split('-', expand=True) #Müşteri kimlik bilgisi kısmını bölüyor
                df_mys['Abone'] = df_mys['Abone'].str.replace(" ","")
                df_mys['Abone']=df_mys['Abone'].astype(float) #abone numaralarını mebbis dosyasıyla eşlemek için tam sayı haline getiriyor
                #gereksiz sütunları temizliyor
                df_mys = df_mys.drop(columns=['Harcama Birimi',
                                            'Fatura Tarihi',
                                            'Boş',
                                            'Müşteri Kimlik Bilgisi',
                                            'Hizmet No',
                                            'Okul'])
                df_mys = df_mys[['Fatura No','Ödenecek Tutar','Tarih','Abone','VKN']]

                df_mebbis = pd.read_excel(yolFatura, header=0) #mebbis dosyasından verileri çekiyor
                cols_to_keep_meb = [
                    'KURUM ADI',
                    'ABONE NUMARASI', 
                    'VERGİ NO', 
                    'FATURA NUMARASI', 
                    'FATURA TARİHİ',
                    'İCMAL NO',
                    'TÜKETİM MİKTARI',
                    'FATURA TUTARI'
                ]
                df_mebbis = df_mebbis.reindex(cols_to_keep_meb, axis=1)
                df_mebbis = df_mebbis.drop(df_mebbis.index[-1]) #tablodaki son satırı siliyor
                df_mebbis['ABONE NUMARASI']=df_mebbis['ABONE NUMARASI'].astype(float) #abone numaralarını mys dosyasıyla eşlemek için tam sayı haline getiriyor
                df_mebbis['VERGİ NO']=df_mebbis['VERGİ NO'].astype(str) #ilerde hata almamak için kolonun veri tipini değiştiriyor
                df_mebbis['İCMAL NO']=df_mebbis['İCMAL NO'].astype(str) #böylece normalde sayısal veriler kayıpsız şekilde metin olarak saklanabiliyor
                
                df_mebbis_dummy = pd.read_excel(yolFatura, header=0, dtype=str) #aynı mebbis dosyasını okuyor ancak vergi no ve icmal metin olarak saklanmalı
                cols_to_keep_dummy = ['VERGİ NO','İCMAL NO'] #bu veriler sayısal olarak saklanırsa başında 0 varsa siliyor
                df_mebbis_dummy = df_mebbis_dummy.reindex(cols_to_keep_dummy, axis=1)
                df_mebbis_dummy = df_mebbis_dummy.drop(df_mebbis_dummy.index[-1])    

                df_kurum = pd.read_excel(yolKurum, dtype=str, header=0) #kurumlar dosyasından verileri çekiyor
                cols_to_keep_kurum = [
                            'VERGİ KİMLİK NO', 
                            'KURUM TÜRÜ'
                        ]
                df_kurum = df_kurum.reindex(cols_to_keep_kurum, axis=1)
                df_kurum =df_kurum.drop(df_kurum.index[-1])
                df_kurum_ana = df_kurum[df_kurum['KURUM TÜRÜ']=='Okul Öncesi'] #anaokullarını ayırıyor
                
                #Tek kaynak formu işlemleri
                toplam = "{:,.2f}".format(df_mebbis['FATURA TUTARI'].sum())
                toplam = toplam.replace(',','-')
                toplam = toplam.replace('.',',')
                toplam = toplam.replace('-','.')
                top2 = df_mebbis['FATURA TUTARI'].sum()
                tutar = f'{toplam} TL ({ParaCevir(top2)})'
                unik = len(df_mebbis['KURUM ADI'].unique())
                if unik<4:
                    kurumlar = ", ".join(str(element) for element in df_mebbis['KURUM ADI'].unique())
                else:
                    kurumlar = f'{len(df_mebbis)} Adet Fatura'
                icmaller = ", ".join(str(element) for element in df_mebbis_dummy['İCMAL NO'].unique())
                ihtiyaç = f'TEMEL EĞİTİM OKULLARI {self.dönem} DÖNEM TELEFON ÖDEMESİ {kurumlar} (İcmal No: {icmaller})'
                tekKaynak = {
                    'firma' : df_firma.values[0][1],
                    'tebligat' : df_firma.values[1][1],
                    'vergi' : df_firma.values[2][1],
                    'telefon' : df_firma.values[3][1],
                    'eposta' : df_firma.values[4][1],
                    'tutar' : tutar,
                    'ihtiyaç' : ihtiyaç,
                    'harcama' : imzaListe[2],
                    'unvan' : imzaListe[3]
                }
                
                tür = 'telefon'
                ilkTertip = f'40.149.423.292.13.68.01.03.05'
                anaTertip = f'40.149.422.291.13.68.01.03.05'
                nitelik = f'Temel Eğitim Okulları {self.dönem} Dönem Telefon Ödemesi' ###
                metin = f'      Müdürlüğümüz  Temel Eğitim Okullarının ({kurumlar}) {self.dönem} dönem {tür} aboneliklerine ait toplam {tutar} borç ödemesi hususunu Onaylarınıza arz ederim.'
                
                harcamaTalimatı = {
                    'tarih' : date.today().strftime("%d.%m.%Y"),
                    'tanım' : faturaTür,
                    'nitelik' : nitelik,
                    'miktar' : tutar,
                    'ödenek1' : '',
                    'ödenek2' : '',
                    'metin' : metin,
                    'ilkTertip': ilkTertip,
                    'anaTertip': anaTertip
                }       


                MYS(df_mys,df_mebbis,df_mebbis_dummy, df_kurum_ana, firma, faturaTür, imzaListe, dosyaYolu, tekKaynak, harcamaTalimatı).MYS()

    def Doğalgaz(self):
        yolFatura = self.yolFatura_D
        yolMYS = self.yolMYS_D
        yolKurum = self.yolKurum
        yolİmza = self.yolİmza

        df_firma = pd.read_excel(io='firma.xlsx',sheet_name='Gaz', header=0)

        firma = df_firma.values[0][1]
        faturaTür = 'Doğalgaz Aboneliği Ödemesi'

        warnings.simplefilter(action='ignore', category=UserWarning)
        
        if os.path.exists(yolFatura):
            if os.path.exists(yolMYS):
                df_imza = pd.read_excel(yolİmza, header=0) #mys'den indirilen dosyayı okuyor
                imzaListe = [
                    df_imza.values[0][0],
                    df_imza.values[1][0],
                    df_imza.values[0][1],
                    df_imza.values[1][1]
                ]

                df_mys = pd.read_excel(yolMYS, header=0) #mys'den indirilen dosyayı okuyor
                cols_to_keep_mys = [
                    'Fatura No',
                    'Harcama Birimi',
                    'Fatura Tarihi',
                    'Ödenecek Tutar',
                    'Müşteri Kimlik Bilgisi'
                ]
                df_mys = df_mys.reindex(cols_to_keep_mys, axis=1) #mys dosyasından verileri çekiyor
                df_mys[['VKN', 'Okul']] = df_mys['Harcama Birimi'].str.split('-', expand=True) #Harcama Birimi kısmını VKN ve Okul olarak ayırıyor
                df_mys[['Tarih', 'Boş']] = df_mys['Fatura Tarihi'].str.split(' ', expand=True) #Fatura tarihi kısımını ayırıyor
                df_mys['Tarih'] = df_mys['Tarih'].str.replace('-', '/') #Fatura tarihini mebbis'ten alınan dosyanın formatına dönüştürüyor
                df_mys['Abone']=df_mys['VKN'] #Müşteri kimlik bilgisi kısmını bölüyor
                
                df_mys['Abone']=df_mys['Abone'].astype(float) #abone numaralarını mebbis dosyasıyla eşlemek için tam sayı haline getiriyor
                #gereksiz sütunları temizliyor
                df_mys = df_mys.drop(columns=['Harcama Birimi',
                                            'Fatura Tarihi',
                                            'Boş',
                                            'Müşteri Kimlik Bilgisi',
                                            'Okul'])
                df_mys = df_mys[['Fatura No','Ödenecek Tutar','Tarih','Abone','VKN']]

                df_mebbis = pd.read_excel(yolFatura, header=0) #mebbis dosyasından verileri çekiyor
                cols_to_keep_meb = [
                    'KURUM ADI',
                    'ABONE NUMARASI', 
                    'VERGİ NO', 
                    'FATURA NUMARASI', 
                    'FATURA TARİHİ',
                    'İCMAL NO',
                    'TÜKETİM MİKTARI',
                    'FATURA TUTARI'
                ]
                df_mebbis = df_mebbis.reindex(cols_to_keep_meb, axis=1)
                df_mebbis = df_mebbis.drop(df_mebbis.index[-1]) #tablodaki son satırı siliyor
                df_mebbis['ABONE NUMARASI']=df_mebbis['VERGİ NO'].astype(int) #abone numaralarını mys dosyasıyla eşlemek için tam sayı haline getiriyor
                df_mebbis['VERGİ NO']=df_mebbis['VERGİ NO'].astype(str) #ilerde hata almamak için kolonun veri tipini değiştiriyor
                df_mebbis['İCMAL NO']=df_mebbis['İCMAL NO'].astype(str) #böylece normalde sayısal veriler kayıpsız şekilde metin olarak saklanabiliyor
                
                df_mebbis_dummy = pd.read_excel(yolFatura, header=0, dtype=str) #aynı mebbis dosyasını okuyor ancak vergi no ve icmal metin olarak saklanmalı
                cols_to_keep_dummy = ['VERGİ NO','İCMAL NO'] #bu veriler sayısal olarak saklanırsa başında 0 varsa siliyor
                df_mebbis_dummy = df_mebbis_dummy.reindex(cols_to_keep_dummy, axis=1)
                df_mebbis_dummy = df_mebbis_dummy.drop(df_mebbis_dummy.index[-1])    

                df_kurum = pd.read_excel(yolKurum, dtype=str, header=0) #kurumlar dosyasından verileri çekiyor
                cols_to_keep_kurum = [
                            'VERGİ KİMLİK NO', 
                            'KURUM TÜRÜ'
                        ]
                df_kurum = df_kurum.reindex(cols_to_keep_kurum, axis=1)
                df_kurum =df_kurum.drop(df_kurum.index[-1])
                df_kurum_ana = df_kurum[df_kurum['KURUM TÜRÜ']=='Okul Öncesi'] #anaokullarını ayırıyor
                
                #Tek kaynak formu işlemleri
                toplam = "{:,.2f}".format(df_mebbis['FATURA TUTARI'].sum())
                toplam = toplam.replace(',','-')
                toplam = toplam.replace('.',',')
                toplam = toplam.replace('-','.')
                top2 = df_mebbis['FATURA TUTARI'].sum()
                tutar = f'{toplam} TL ({ParaCevir(top2)})'
                unik = len(df_mebbis['KURUM ADI'].unique())
                if unik<4:
                    kurumlar = ", ".join(str(element) for element in df_mebbis['KURUM ADI'].unique())
                else:
                    kurumlar = f'{len(df_mebbis)} Adet Fatura'
                icmaller = ", ".join(str(element) for element in df_mebbis_dummy['İCMAL NO'].unique())
                ihtiyaç = f'TEMEL EĞİTİM OKULLARI {self.dönem} DÖNEM TELEFON ÖDEMESİ {kurumlar} (İcmal No: {icmaller})'
                tekKaynak = {
                    'firma' : df_firma.values[0][1],
                    'tebligat' : df_firma.values[1][1],
                    'vergi' : df_firma.values[2][1],
                    'telefon' : df_firma.values[3][1],
                    'eposta' : df_firma.values[4][1],
                    'tutar' : tutar,
                    'ihtiyaç' : ihtiyaç,
                    'harcama' : imzaListe[2],
                    'unvan' : imzaListe[3]
                }
                
                tür = 'doğalgaz'
                ilkTertip = f'40.149.423.18734.13.68.01.03.02'
                anaTertip = f'40.149.422.18735.13.68.01.03.02'
                nitelik = f'Temel Eğitim Okulları {self.dönem} Dönem Doğalgaz Ödemesi' ###
                metin = f'      Müdürlüğümüz  Temel Eğitim Okullarının ({kurumlar}) {self.dönem} dönem {tür} aboneliklerine ait toplam {tutar} borç ödemesi hususunu Onaylarınıza arz ederim.'
                
                harcamaTalimatı = {
                    'tarih' : date.today().strftime("%d.%m.%Y"),
                    'tanım' : faturaTür,
                    'nitelik' : nitelik,
                    'miktar' : tutar,
                    'ödenek1' : '',
                    'ödenek2' : '',
                    'metin' : metin,
                    'ilkTertip': ilkTertip,
                    'anaTertip': anaTertip
                }       


                MYS(df_mys,df_mebbis,df_mebbis_dummy, df_kurum_ana, firma, faturaTür, imzaListe, dosyaYolu, tekKaynak, harcamaTalimatı).MYS()


class MYS():
    def __init__(self,
                 df_mys:pd.DataFrame,
                 df_mebbis:pd.DataFrame,
                 df_mebbis_dummy:pd.DataFrame,
                 df_kurum_ana:pd.DataFrame,
                 firma:str,
                 faturaTür:str,
                 imzaListe:list,
                 dosyaYolu:str,
                 tekKaynak: dict,
                 harcamaTalimatı: dict
                 ):
        
        self.df_ana = pd.DataFrame() #boş dataframe ler oluşturuyor, düzeltilmiş dataframe i ayırıp birleştirecez
        self.df_ilk = pd.DataFrame()

        self.df_mys = df_mys
        self.df_mebbis = df_mebbis
        self.df_mebbis_dummy = df_mebbis_dummy
        self.df_kurum_ana = df_kurum_ana
        self.firma = firma
        self.faturaTür = faturaTür
        self.imzaListe = imzaListe
        self.dosyaYolu = dosyaYolu
        self.tekKaynak = tekKaynak
        self.harcamaTalimatı = harcamaTalimatı

        self.meb_kur_col = df_mebbis.columns.get_loc('KURUM ADI')
        self.meb_abo_col = df_mebbis.columns.get_loc('ABONE NUMARASI')
        self.meb_ver_col = df_mebbis.columns.get_loc('VERGİ NO')
        self.meb_fat_col = df_mebbis.columns.get_loc('FATURA NUMARASI')
        self.meb_tar_col = df_mebbis.columns.get_loc('FATURA TARİHİ')
        self.meb_icm_col = df_mebbis.columns.get_loc('İCMAL NO')
        self.meb_tuk_col = df_mebbis.columns.get_loc('TÜKETİM MİKTARI')
        self.meb_tut_col = df_mebbis.columns.get_loc('FATURA TUTARI')
        
        self.mebD_ver_col = df_mebbis_dummy.columns.get_loc('VERGİ NO')
        self.mebD_icm_col = df_mebbis_dummy.columns.get_loc('İCMAL NO')
        
        self.mys_fat_col = df_mys.columns.get_loc('Fatura No')
        self.mys_tut_col = df_mys.columns.get_loc('Ödenecek Tutar')
        self.mys_tar_col = df_mys.columns.get_loc('Tarih')
        self.mys_ver_col = df_mys.columns.get_loc('VKN')

        yolTemp = r'%s\Taslaklar\temp.xlsx'%(dosyaYolu)
        
        self.wb = load_workbook(yolTemp)
        self.ws = self.wb['Liste']
        
    def MYS(self):
        wb = self.wb
        ws = self.ws
        tekKaynak = self.tekKaynak
        harcamaTalimatı = self.harcamaTalimatı

        for i in range(len(self.df_mebbis.index)): #baz olarak mebbis dosyasını alıp oradaki tüm satırları teker teker işleme alıyor
            indices_fatura = np.where(self.df_mys['Fatura No'] == self.df_mebbis.iat[i,self.meb_fat_col]) #iki dosyadaki fatura numaralarını eşleştiriyor
            row_indices_fatura = indices_fatura[0] #faturaların mys dosyasındaki konumunu buluyor
            if row_indices_fatura.size == 0: #Fatura numarasının mys dosyasında olup olmadığını kontrol ediyor
                aboneNo = self.df_mys['Abone'].values == self.df_mebbis.iat[i,self.meb_abo_col] #Fatura numarası mys dosyasında yoksa abone numarasını buluyor.
                if max(aboneNo)==True: #abone numarasının mys dosyasında olup olmadığını kontrol ediyor
                    row_indices_df_mys = np.where(aboneNo == True) #Abone numarasının hangi sırada olduğunu buluyor
                    #Tutar ve fatura tarihi kontrolü
                    if self.df_mebbis.iat[i,self.meb_tut_col] == self.df_mys.iat[row_indices_df_mys[0][0],self.mys_tut_col] and self.df_mebbis.iat[i,self.meb_tar_col] == self.df_mys.iat[row_indices_df_mys[0][0],self.mys_tar_col]:
                        #Tutar ve tarih iki tabloda da aynı ise mys'deki fatura numarasını alıyor
                        self.df_mebbis.iat[i,self.meb_fat_col] = self.df_mys.iat[row_indices_df_mys[0][0],self.mys_fat_col]                          
                    else: #tutar ve tarih tutmuyorsa fatura girişi hatalı demektir 
                        print(f'{self.df_mebbis.iat[i,self.meb_fat_col]} Fatura girişi hatalı!')
                else: #abone numarası yoksa zaten mys'de yok demektir
                    print(f'{self.df_mebbis.iat[i,self.meb_fat_col]} MYS de yok!')                      
            else:
                #fatura var ancak tutarlar farklı ise tutarı mys'den çekiyor
                if self.df_mebbis.iat[i, self.meb_tut_col] != self.df_mys.iat[row_indices_fatura[0],self.mys_tut_col]:
                    self.df_mebbis.iat[i, self.meb_tut_col] = self.df_mys.iat[row_indices_fatura[0],self.mys_tut_col]
                    #tutar değişimi olursa uyarı veriyor, hangi faturanın tutarı nasıl değişmiş görebiliyoruz
                    print(f'{self.df_mebbis.iat[i, self.meb_fat_col]} numaralı fatura tutarı {self.df_mebbis.iat[i, self.eb_tut_col]} iken {self.df_mys.iat[row_indices_fatura[0],self.mys_tut_col]} oldu.')
            #vergi numarasını metin haline getiriyor böylece başta 0 varsa yoksaymıyor
            self.df_mebbis.iat[i,self.meb_ver_col] = self.df_mebbis_dummy.iat[i,self.mebD_ver_col]
            self.df_mebbis.iat[i,self.meb_icm_col] = self.df_mebbis_dummy.iat[i,self.mebD_icm_col]
            #kurum anaokulları içerisinde mi onu kontrol ediyor 
            if self.df_kurum_ana.isin([self.df_mebbis.iat[i,self.meb_ver_col]]).any().any():
                #anaokulları içerisindeyse anaokulu dataframe ine ekliyor
                self.df_ana = pd.concat([self.df_ana,self.df_mebbis.iloc[[i]]])
            else:
                #değilse ilkokullar dataframe ine ekliyor
                self.df_ilk = pd.concat([self.df_ilk,self.df_mebbis.iloc[[i]]])

        ws.cell(row=2,column=2).value = self.faturaTür
        Hizala.Solda(self.wb,2,2)
        ws.cell(row=3,column=2).value = self.firma
        Hizala.Solda(wb,3,2)

        başlangıç = 5

        if self.df_ilk.empty:
            toplam_ilk = 0
            if self.df_ana.empty:
                toplam_ana = 0
                pass
            else:
                toplam_ana = self.df_ana['FATURA TUTARI'].sum()
                self.işle(wb,self.df_ana,başlangıç)
                başlangıç += len(self.df_ana)
                ws['L6'] = f'=SUM(I{başlangıç-len(self.df_ana)}:I{başlangıç-1})' ####
                ws.cell(row=başlangıç, column=1).value = 'ANAOKULLARI TOPLAMI'
                Hizala.SağdaKalın(wb,başlangıç,1)
                ws.merge_cells(f'A{başlangıç}:G{başlangıç}')
                ws.cell(row=başlangıç, column=self.meb_tut_col+1).value = toplam_ana
                Hizala.SağdaKalın(wb,başlangıç,self.meb_tut_col+1)   
                ws.cell(row=başlangıç, column=self.meb_tut_col+1).number_format = '#,##0.00'
                ws['M6'] = f'=H{başlangıç}-L6' ###     
                başlangıç += 1
        else:
            toplam_ilk = self.df_ilk['FATURA TUTARI'].sum()
            self.işle(wb, self.df_ilk,başlangıç)
            başlangıç += len(self.df_ilk)
            ws['L5'] = f'=SUM(I5:I{başlangıç-1})' ####
            
            ws.cell(row=başlangıç, column=1).value = 'İLKÖĞRETİM TOPLAMI'
            Hizala.SağdaKalın(wb,başlangıç,1)
            ws.merge_cells(f'A{başlangıç}:G{başlangıç}')
            ws.cell(row=başlangıç, column=self.meb_tut_col+1).value = toplam_ilk
            Hizala.SağdaKalın(wb,başlangıç,self.meb_tut_col+1)
            ws.cell(row=başlangıç, column=self.meb_tut_col+1).number_format = '#,##0.00'
            ws['M5'] = f'=H{başlangıç}-L5' ###
            başlangıç += 1
            
            if self.df_ana.empty:
                toplam_ana = 0
            else:
                toplam_ana = self.df_ana['FATURA TUTARI'].sum()
                self.işle(wb,self.df_ana,başlangıç)
                başlangıç +=len(self.df_ana)
                ws['L6'] = f'=SUM(I{başlangıç-len(self.df_ana)}:I{başlangıç-1})' ####
                ws.cell(row=başlangıç, column=1).value = 'ANAOKULLARI TOPLAMI'
                Hizala.SağdaKalın(wb,başlangıç,1)
                ws.merge_cells(f'A{başlangıç}:G{başlangıç}')
                ws.cell(row=başlangıç, column=self.meb_tut_col+1).value = toplam_ana
                Hizala.SağdaKalın(wb,başlangıç,self.meb_tut_col+1)   
                ws.cell(row=başlangıç, column=self.meb_tut_col+1).number_format = '#,##0.00'
                ws['M6'] = f'=H{başlangıç}-L6' ###             
                başlangıç +=1
                
        ws.cell(row=başlangıç, column=1).value = 'GENEL TOPLAM'
        Hizala.SağdaKalın(wb,başlangıç,1)
        ws.merge_cells(f'A{başlangıç}:G{başlangıç}')      
        ws.cell(row=başlangıç, column=self.meb_tut_col+1).value = toplam_ilk + toplam_ana
        Hizala.SağdaKalın(wb,başlangıç,self.meb_tut_col+1)   
        ws.cell(row=başlangıç, column=self.meb_tut_col+1).number_format = '#,##0.00'   
        başlangıç += 4

        ws.cell(row=başlangıç, column=1).value = self.imzaListe[0]
        Hizala.SadeOrta(wb,başlangıç,1)
        ws.cell(row=başlangıç+1, column=1).value = self.imzaListe[1]
        Hizala.SadeOrta(wb,başlangıç+1,1)
        ws.cell(row=başlangıç, column=7).value = self.imzaListe[2]
        Hizala.SadeOrta(wb,başlangıç,7)
        ws.cell(row=başlangıç+1, column=7).value = self.imzaListe[3]
        Hizala.SadeOrta(wb,başlangıç+1,7)

        ws.print_area = f'A1:H{başlangıç+3}'

        wb.save(r'%s\Fatura_%s.xlsx'%(desktop,self.faturaTür))

        wbTek = load_workbook(r'%s\Fatura_%s.xlsx'%(desktop,self.faturaTür))
        wsTek = wbTek['Tek']

        wsTek['C5'] = tekKaynak['firma']
        wsTek['C6'] = tekKaynak['tebligat']
        wsTek['C7'] = tekKaynak['vergi']
        wsTek['C8'] = tekKaynak['telefon']
        wsTek['C9'] = tekKaynak['eposta']
        wsTek['C12'] = tekKaynak['tutar']
        wsTek['A15'] = tekKaynak['ihtiyaç']
        wsTek['C40'] = tekKaynak['harcama']
        wsTek['C41'] = tekKaynak['unvan']

        wbTek.save(r'%s\Fatura_%s.xlsx'%(desktop,self.faturaTür))

        wbHar = load_workbook(r'%s\Fatura_%s.xlsx'%(desktop,self.faturaTür))
        wsHar = wbHar['Harcama']
    
        ilk = toplam_ilk
        ana = toplam_ana
        
        ilkT = '{:,.2f}'.format(ilk)
        ilkT = ilkT.replace(',','-')
        ilkT = ilkT.replace('.',',')
        ilkT = ilkT.replace('-','.')        

        anaT = '{:,.2f}'.format(ana)
        anaT = anaT.replace(',','-')
        anaT = anaT.replace('.',',')
        anaT = anaT.replace('-','.')
        
        ilkTertip = harcamaTalimatı['ilkTertip']
        anaTertip = harcamaTalimatı['anaTertip']
        
        ilkS = f'{ilkTertip} ({ilkT} TL)'
        anaS = f'{anaTertip} ({anaT} TL)'
        
        if ilk == 0:
            ilkS = anaS
            anaS = ''
        if ana == 0:
            anaS = ''
            
        self.harcamaTalimatı['ödenek1'] = ilkS
        self.harcamaTalimatı['ödenek2'] = anaS
        
        wsHar['C4'] = harcamaTalimatı['tarih']
        wsHar['C8'] = harcamaTalimatı['tanım']
        wsHar['C9'] = harcamaTalimatı['nitelik']
        wsHar['C10'] = harcamaTalimatı['miktar']
        wsHar['C11'] = harcamaTalimatı['ödenek1']
        wsHar['C12'] = harcamaTalimatı['ödenek2']
        wsHar['A16'] = harcamaTalimatı['metin']
        wsHar['A23'] = self.imzaListe[0]
        wsHar['A24'] = self.imzaListe[1]
        wsHar['D23'] = self.imzaListe[2]
        wsHar['D24'] = self.imzaListe[3]  
        
        wbHar.save(r'%s\Fatura_%s.xlsx'%(desktop,self.faturaTür))    

    def işle(self, wb:Workbook, df:pd.DataFrame, başlangıç:int):
        ws = wb.active
        for i in range(len(df.values)):
            ws.cell(row=başlangıç+i, column=self.meb_kur_col+1).value = df.values[i][self.meb_kur_col]
            Hizala.Solda(wb,başlangıç+i,self.meb_kur_col+1)
            ws.cell(row=başlangıç+i, column=self.meb_abo_col+1).value = df.values[i][self.meb_abo_col]
            Hizala.Ortala(wb,başlangıç+i,self.meb_abo_col+1)
            ws.cell(row=başlangıç+i, column=self.meb_ver_col+1).value = df.values[i][self.meb_ver_col]
            Hizala.Ortala(wb,başlangıç+i,self.meb_ver_col+1)
            ws.cell(row=başlangıç+i, column=self.meb_fat_col+1).value = df.values[i][self.meb_fat_col]
            Hizala.Ortala(wb,başlangıç+i,self.meb_fat_col+1)
            ws.cell(row=başlangıç+i, column=self.meb_tar_col+1).value = df.values[i][self.meb_tar_col]
            Hizala.Ortala(wb,başlangıç+i,self.meb_tar_col+1)
            ws.cell(row=başlangıç+i, column=self.meb_icm_col+1).value = df.values[i][self.meb_icm_col]
            Hizala.Ortala(wb,başlangıç+i,self.meb_icm_col+1)
            ws.cell(row=başlangıç+i, column=self.meb_tuk_col+1).value = df.values[i][self.meb_tuk_col]
            Hizala.Ortala(wb,başlangıç+i,self.meb_tuk_col+1)
            ws.cell(row=başlangıç+i, column=self.meb_tut_col+1).value = df.values[i][self.meb_tut_col]
            ws.cell(row=başlangıç+i, column=self.meb_tut_col+1).number_format = '#,##0.00'
            Hizala.Sağda(wb,başlangıç+i,self.meb_tut_col+1)      
 
class Hizala():
    def Solda(wb:Workbook, row, column):
        ws = wb.active
        ft = Font(color='FF000000', name='Times New Roman', size=10)
        thin = Side(border_style="thin", color="FF000000")
        
        ws.cell(row=row, column=column).font = ft
        ws.cell(row=row, column=column).alignment = Alignment(horizontal="left", vertical="center")
        ws.cell(row=row, column=column).border = Border(top=thin, left=thin, right=thin, bottom=thin)

    def Ortala(wb:Workbook, row, column):
        ws = wb.active
        ft = Font(color='FF000000', name='Times New Roman', size=10)
        thin = Side(border_style="thin", color="FF000000")
        
        ws.cell(row=row, column=column).font = ft
        ws.cell(row=row, column=column).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=row, column=column).border = Border(top=thin, left=thin, right=thin, bottom=thin)

    def SadeOrta(wb:Workbook, row, column):
        ws = wb.active
        ft = Font(color='FF000000', name='Times New Roman', size=10)
        
        ws.cell(row=row, column=column).font = ft
        ws.cell(row=row, column=column).alignment = Alignment(horizontal="center", vertical="center")

    def Sağda(wb:Workbook, row, column):
        ws = wb.active
        ft = Font(color='FF000000', name='Times New Roman', size=10)
        thin = Side(border_style="thin", color="FF000000")
        
        ws.cell(row=row, column=column).font = ft
        ws.cell(row=row, column=column).alignment = Alignment(horizontal="right", vertical="center")
        ws.cell(row=row, column=column).border = Border(top=thin, left=thin, right=thin, bottom=thin)

    def SağdaKalın(wb:Workbook, row, column):
        ws = wb.active
        ft = Font(color='FF000000', name='Times New Roman', size=12, bold=True)
        thin = Side(border_style="thin", color="FF000000")
        
        ws.cell(row=row, column=column).font = ft
        ws.cell(row=row, column=column).alignment = Alignment(horizontal="right", vertical="center")
        ws.cell(row=row, column=column).border = Border(top=thin, left=thin, right=thin, bottom=thin)

