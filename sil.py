from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Border, Side, Alignment
import pandas as pd
import numpy as np
import os

f = open("dosyaYolu.txt", "r")
dosyaYolu = r'%s'%(f.read())
f.close()
yolFatura = r'%s\Su\Fatura Listesi.xlsx'%(dosyaYolu)
yolTakip = r'%s\Taslaklar\suTakip.xlsx'%(dosyaYolu) #takip dosyasına ulaşıyor

wb = load_workbook(yolTakip) #takip excelini açıyor
ws = wb['2024'] #çalışma sayfasına gidiyor burada yıl kısmını GUI deki dönemden çekeceğiz

col = 7 #GUI deki dönemden çekeceğiz

#ws.cell(row=2,column=col+4).value = float(20) #takip exceline veri işleme

#wb.save(yolTakip) #veri işlenen takip excelini kaydetme"""

df_mebbis = pd.read_excel(yolFatura, header=0)
cols_to_keep_meb = [
                    'KURUM ADI',
                    'ABONE NUMARASI', 
                    'VERGİ NO', 
                    'FATURA NUMARASI', 
                    'FATURA TARİHİ',
                    'İCMAL NO',
                    'TÜKETİM MİKTARI',
                    'FATURA TUTARI'
                ]
df_mebbis = df_mebbis.reindex(cols_to_keep_meb, axis=1)
df_mebbis = df_mebbis.drop(df_mebbis.index[-1]) #tablodaki son satırı siliyor

#Aşağıdaki kısmı düzenlemeye ihtiyaç olabilir, veriler istediğim türde gelmedi 

df_mebbis['ABONE NUMARASI']=df_mebbis['ABONE NUMARASI'].astype(int) #abone numaralarını mys dosyasıyla eşlemek için tam sayı haline getiriyor
df_mebbis['VERGİ NO']=df_mebbis['VERGİ NO'].astype(str) #ilerde hata almamak için kolonun veri tipini değiştiriyor
df_mebbis['İCMAL NO']=df_mebbis['İCMAL NO'].astype(str) #böylece normalde sayısal veriler kayıpsız şekilde metin olarak saklanabiliyor


meb_abo_col = df_mebbis.columns.get_loc('ABONE NUMARASI')
meb_tuk_col = df_mebbis.columns.get_loc('TÜKETİM MİKTARI')


df_takip = pd.read_excel(yolTakip, header=0) #su takip dosyasındaki verileri çekiyor. sadece abone numarası ile son sütununa ihtiyaç var
df_takip = df_takip.fillna(0) #boş hücrelerin yerine 0 yazıyor, gerekli olmayabilir çünkü veriyi yeniden kaydedilmiş excelden çekecez

for i in range(len(df_mebbis.index)):
    indices_fatura = np.where(df_takip['ABONE NUMARASI'] == df_mebbis.iat[i,meb_abo_col])
    row_indices_fatura = indices_fatura[0]
    ws.cell(row=int(row_indices_fatura[0])+2,column=col+4).value = df_mebbis.iat[i,meb_tuk_col]

wb.save(yolTakip) #veri işlenen takip excelini kaydetme"""

#print(df_mebbis)
